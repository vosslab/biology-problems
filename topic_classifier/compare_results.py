#!/usr/bin/env python3

"""Compare topic classifier results across multiple LLM model runs.

Auto-discovers results-*/ directories and reports where models agree or
disagree on chapter and topic assignments. Outputs a rich terminal summary
and an XLSX workbook with colored disagreement tables.
"""

# Standard Library
import os
import csv
import glob
import argparse

# PIP3 modules
import openpyxl
import openpyxl.styles
import openpyxl.utils
import rich.console
import rich.table

# local repo modules
import topic_classifier.index_parser_lib as index_parser

console = rich.console.Console(highlight=False)

#============================================
def parse_args() -> argparse.Namespace:
	"""Parse command-line arguments."""
	parser = argparse.ArgumentParser(
		description="Compare topic classifier results across multiple model runs.",
	)
	parser.add_argument(
		'-b', '--base-dir', dest='base_dir', type=str,
		default=None,
		help="Base directory containing results-*/ folders (default: script directory)",
	)
	parser.add_argument(
		'-o', '--output', dest='output_xlsx', type=str,
		default=None,
		help="Output XLSX path (default: output/comparison_results.xlsx)",
	)
	parser.add_argument(
		'-c', '--csv-output-dir', dest='csv_output_dir', type=str,
		default=None,
		help="Override output dir for agreed_tasks CSVs (default: ./agreed_tasks/ in CWD)",
	)
	args = parser.parse_args()
	return args

#============================================
def load_results_dir(dir_path: str) -> dict:
	"""Read all *_tasks.csv files from one results directory.

	Args:
		dir_path: path to a results-*/ directory

	Returns:
		dict mapping script path to (chapter, topic) tuple
	"""
	assignments = {}
	# Skip non-classification files
	skip_files = {"llm_errors.csv", "no_bbq_file.csv"}
	csv_pattern = os.path.join(dir_path, "*_tasks.csv")
	for csv_path in sorted(glob.glob(csv_pattern)):
		basename = os.path.basename(csv_path)
		if basename in skip_files:
			continue
		with open(csv_path, "r") as f:
			reader = csv.reader(f)
			# Skip header row
			next(reader)
			for row in reader:
				if not row or all(cell.strip() == "" for cell in row):
					continue
				# Pad short rows
				while len(row) < 3:
					row.append("")
				chapter = row[0].strip()
				topic = row[1].strip()
				script = row[2].strip()
				if not script:
					continue
				assignments[script] = (chapter, topic)
	return assignments

#============================================
def load_all_results(base_dir: str) -> dict:
	"""Discover and load all results-*/ directories.

	Args:
		base_dir: directory containing results-*/ folders

	Returns:
		dict mapping model name to {script: (chapter, topic)} dict
	"""
	all_results = {}
	pattern = os.path.join(base_dir, "results-*/")
	for dir_path in sorted(glob.glob(pattern)):
		if not os.path.isdir(dir_path):
			continue
		# Extract model name from directory name
		dir_name = os.path.basename(dir_path.rstrip("/"))
		model_name = dir_name.replace("results-", "")
		assignments = load_results_dir(dir_path)
		if assignments:
			all_results[model_name] = assignments
			console.print(f"  Loaded [bold]{model_name}[/bold]: {len(assignments)} scripts")
	return all_results

#============================================
def get_all_scripts(all_results: dict) -> set:
	"""Collect all unique script paths across all models.

	Args:
		all_results: output of load_all_results()

	Returns:
		set of script paths
	"""
	scripts = set()
	for assignments in all_results.values():
		scripts.update(assignments.keys())
	return scripts

#============================================
def load_known_overlaps(csv_path: str) -> dict:
	"""Read known_overlaps.csv into a lookup dict.

	Args:
		csv_path: path to known_overlaps.csv

	Returns:
		dict mapping script path to set of (chapter, topic) tuples
	"""
	overlaps = {}
	if not os.path.isfile(csv_path):
		return overlaps
	with open(csv_path, "r") as f:
		reader = csv.DictReader(f)
		for row in reader:
			script = row["script"].strip()
			chapter = row["chapter"].strip()
			topic = row["topic"].strip()
			if script not in overlaps:
				overlaps[script] = set()
			overlaps[script].add((chapter, topic))
	return overlaps

#============================================
def find_disagreements(all_results: dict, known_overlaps: dict = None) -> dict:
	"""Compare classifications across models.

	Args:
		all_results: output of load_all_results()
		known_overlaps: output of load_known_overlaps(), or None

	Returns:
		dict with keys: chapter_agree, topic_agree, chapter_disagree,
		topic_disagree, unique_scripts, known_overlap. Each is a list.
	"""
	if known_overlaps is None:
		known_overlaps = {}
	models = sorted(all_results.keys())
	all_scripts = sorted(get_all_scripts(all_results))

	chapter_agree = []
	topic_agree = []
	chapter_disagree = []
	topic_disagree = []
	unique_scripts = []
	known_overlap = []

	for script in all_scripts:
		# Collect assignments per model for this script
		present_in = {}
		for model in models:
			if script in all_results[model]:
				present_in[model] = all_results[model][script]

		# Script not in all models
		if len(present_in) < len(models):
			unique_scripts.append({
				"script": script,
				"assignments": present_in,
				"missing_from": [m for m in models if m not in present_in],
			})
			# Still compare among models that have it
			if len(present_in) < 2:
				continue

		# Get all chapters and topics assigned
		chapters = set(ch for ch, _tp in present_in.values())
		topics = set(tp for _ch, tp in present_in.values())

		if len(chapters) > 1 or len(topics) > 1:
			# Check if all assignments are in known overlaps
			valid_set = known_overlaps.get(script, set())
			all_known = all(
				(ch, tp) in valid_set
				for ch, tp in present_in.values()
			)
			if all_known and valid_set:
				known_overlap.append({
					"script": script,
					"assignments": present_in,
					"valid_assignments": valid_set,
				})
			elif len(chapters) > 1:
				# Chapter-level disagreement
				chapter_disagree.append({
					"script": script,
					"assignments": present_in,
				})
			else:
				# Same chapter, different topic
				topic_disagree.append({
					"script": script,
					"assignments": present_in,
				})
		else:
			# Full agreement - store script and the common assignment
			chapter, topic = list(present_in.values())[0]
			topic_agree.append({
				"script": script,
				"chapter": chapter,
				"topic": topic,
			})
			chapter_agree.append(script)

	results = {
		"chapter_agree": chapter_agree,
		"topic_agree": topic_agree,
		"chapter_disagree": chapter_disagree,
		"topic_disagree": topic_disagree,
		"unique_scripts": unique_scripts,
		"known_overlap": known_overlap,
	}
	return results

#============================================
def print_overview_table(all_results: dict) -> None:
	"""Print per-model chapter distribution table.

	Args:
		all_results: output of load_all_results()
	"""
	models = sorted(all_results.keys())

	# Collect all chapters
	all_chapters = set()
	for assignments in all_results.values():
		for chapter, _topic in assignments.values():
			all_chapters.add(chapter)

	table = rich.table.Table(title="Scripts per Chapter by Model")
	table.add_column("Chapter", style="bold")
	for model in models:
		table.add_column(model, justify="right")

	for chapter in sorted(all_chapters):
		row = [chapter]
		for model in models:
			count = sum(
				1 for ch, _tp in all_results[model].values()
				if ch == chapter
			)
			row.append(str(count))
		table.add_row(*row)

	# Total row
	total_row = ["[bold]TOTAL[/bold]"]
	for model in models:
		total_row.append(f"[bold]{len(all_results[model])}[/bold]")
	table.add_row(*total_row)

	console.print()
	console.print(table)

#============================================
def print_agreement_summary(disagreements: dict, total_scripts: int) -> None:
	"""Print summary counts of agreement and disagreement.

	Args:
		disagreements: output of find_disagreements()
		total_scripts: total unique scripts across all models
	"""
	n_topic_agree = len(disagreements["topic_agree"])
	n_chapter_disagree = len(disagreements["chapter_disagree"])
	n_topic_disagree = len(disagreements["topic_disagree"])
	n_unique = len(disagreements["unique_scripts"])
	n_known = len(disagreements["known_overlap"])

	console.print()
	console.print("[bold]Agreement Summary[/bold]")
	console.print(f"  Total unique scripts: {total_scripts}")
	console.print(f"  Full agreement (chapter + topic): [green]{n_topic_agree}[/green]")
	console.print(f"  Known overlaps (all valid): [blue]{n_known}[/blue]")
	console.print(f"  Chapter disagreements: [red]{n_chapter_disagree}[/red]")
	console.print(f"  Topic disagreements (same chapter): [yellow]{n_topic_disagree}[/yellow]")
	console.print(f"  Not in all models: [cyan]{n_unique}[/cyan]")

#============================================
def print_chapter_disagreements(disagreements: dict, models: list, topic_lookup: dict) -> None:
	"""Print table of scripts with chapter-level disagreements.

	Args:
		disagreements: output of find_disagreements()
		models: sorted list of model names
		topic_lookup: output of build_topic_lookup()
	"""
	items = disagreements["chapter_disagree"]
	if not items:
		console.print("\n[green]No chapter-level disagreements.[/green]")
		return

	table = rich.table.Table(title=f"Chapter Disagreements ({len(items)} scripts)", show_lines=True)
	table.add_column("Script", style="bold")
	for model in models:
		table.add_column(model, min_width=22)

	for item in items:
		# Shorten script path for display
		script_short = _shorten_script_path(item["script"])
		row = [script_short]
		for model in models:
			if model in item["assignments"]:
				chapter, topic = item["assignments"][model]
				label = format_assignment(chapter, topic, topic_lookup)
				row.append(f"[red]{label}[/red]")
			else:
				row.append("[dim]---[/dim]")
		table.add_row(*row)

	console.print()
	console.print(table)

#============================================
def print_topic_disagreements(disagreements: dict, models: list, topic_lookup: dict) -> None:
	"""Print table of scripts with topic-level disagreements.

	Args:
		disagreements: output of find_disagreements()
		models: sorted list of model names
		topic_lookup: output of build_topic_lookup()
	"""
	items = disagreements["topic_disagree"]
	if not items:
		console.print("\n[green]No topic-level disagreements.[/green]")
		return

	table = rich.table.Table(title=f"Topic Disagreements ({len(items)} scripts)", show_lines=True)
	table.add_column("Script", style="bold")
	for model in models:
		table.add_column(model, min_width=22)

	for item in items:
		script_short = _shorten_script_path(item["script"])
		row = [script_short]
		for model in models:
			if model in item["assignments"]:
				chapter, topic = item["assignments"][model]
				topic_name = topic_lookup.get((chapter, topic), "")
				row.append(f"{chapter}/[yellow]{topic} {topic_name}[/yellow]")
			else:
				row.append("[dim]---[/dim]")
		table.add_row(*row)

	console.print()
	console.print(table)

#============================================
def print_unique_scripts(disagreements: dict, models: list, topic_lookup: dict) -> None:
	"""Print scripts that appear in only some models.

	Args:
		disagreements: output of find_disagreements()
		models: sorted list of model names
		topic_lookup: output of build_topic_lookup()
	"""
	items = disagreements["unique_scripts"]
	if not items:
		console.print("\n[green]All scripts present in all models.[/green]")
		return

	table = rich.table.Table(title=f"Scripts Not in All Models ({len(items)} scripts)", show_lines=True)
	table.add_column("Script", style="bold")
	for model in models:
		table.add_column(model, min_width=22)

	for item in items:
		script_short = _shorten_script_path(item["script"])
		row = [script_short]
		for model in models:
			if model in item["assignments"]:
				chapter, topic = item["assignments"][model]
				row.append(format_assignment(chapter, topic, topic_lookup))
			else:
				row.append("[dim]missing[/dim]")
		table.add_row(*row)

	console.print()
	console.print(table)

#============================================
def build_topic_lookup() -> dict:
	"""Build a lookup dict mapping (chapter, topic_id) to topic name.

	Returns:
		dict mapping (chapter, topic_id) tuples to topic name strings,
		e.g. {("biochemistry", "topic01"): "Life Molecules", ...}
	"""
	all_indexes = index_parser.load_all_indexes()
	lookup = {}
	for subject, data in all_indexes.items():
		for topic in data["topics"]:
			key = (subject, topic["topic_id"])
			lookup[key] = topic["name"]
	return lookup

#============================================
def format_assignment(chapter: str, topic: str, topic_lookup: dict) -> str:
	"""Format a chapter/topic assignment with the topic title.

	Args:
		chapter: subject name like 'biochemistry'
		topic: topic id like 'topic07'
		topic_lookup: output of build_topic_lookup()

	Returns:
		formatted string like 'biochemistry/topic07 Enzymes'
	"""
	topic_name = topic_lookup.get((chapter, topic), "")
	if topic_name:
		result = f"{chapter}/{topic} {topic_name}"
	else:
		result = f"{chapter}/{topic}"
	return result

#============================================
def _shorten_script_path(script: str) -> str:
	"""Shorten a script path for CLI display.

	Args:
		script: full script path from CSV

	Returns:
		shortened path with 'problems/' prefix and '-problems' folder suffix removed
	"""
	clean = script
	if clean.startswith("problems/"):
		clean = clean[len("problems/"):]
	clean = clean.replace("-problems/", "/", 1)
	return clean

#============================================
def split_script_path(script: str) -> tuple:
	"""Split a script path into folder and basename.

	Args:
		script: script path like 'inheritance-problems/blood_type_mother.py'

	Returns:
		(folder, basename) tuple
	"""
	# Strip leading 'problems/' prefix and '-problems' folder suffix
	clean = script
	if clean.startswith("problems/"):
		clean = clean[len("problems/"):]
	# Remove the '-problems/' part from folder names like 'inheritance-problems/'
	clean = clean.replace("-problems/", "/", 1)
	folder = os.path.dirname(clean)
	basename = os.path.basename(clean)
	return (folder, basename)

#============================================
def auto_size_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
	"""Auto-size all columns in a worksheet based on content width.

	Args:
		ws: openpyxl worksheet
	"""
	for col_cells in ws.columns:
		max_length = 0
		col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
		for cell in col_cells:
			cell_len = len(str(cell.value)) if cell.value else 0
			if cell_len > max_length:
				max_length = cell_len
		# Add padding
		adjusted_width = min(max_length + 3, 50)
		ws.column_dimensions[col_letter].width = adjusted_width

#============================================
def write_xlsx(all_results: dict, disagreements: dict, output_path: str, topic_lookup: dict) -> None:
	"""Write comparison results to a multi-sheet XLSX workbook.

	Args:
		all_results: output of load_all_results()
		disagreements: output of find_disagreements()
		output_path: path for the output XLSX file
		topic_lookup: output of build_topic_lookup()
	"""
	os.makedirs(os.path.dirname(output_path), exist_ok=True)
	models = sorted(all_results.keys())

	# Define fill colors
	header_fill = openpyxl.styles.PatternFill(start_color="4472C4", fill_type="solid")
	header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
	bold_font = openpyxl.styles.Font(bold=True)
	green_fill = openpyxl.styles.PatternFill(start_color="92D050", fill_type="solid")
	red_fill = openpyxl.styles.PatternFill(start_color="FF6B6B", fill_type="solid")
	yellow_fill = openpyxl.styles.PatternFill(start_color="FFD966", fill_type="solid")
	cyan_fill = openpyxl.styles.PatternFill(start_color="9BC2E6", fill_type="solid")
	gray_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", fill_type="solid")

	wb = openpyxl.Workbook()

	# --- Sheet 1: Overview ---
	ws_overview = wb.active
	ws_overview.title = "Overview"
	# Collect all chapters
	all_chapters = set()
	for assignments in all_results.values():
		for chapter, _topic in assignments.values():
			all_chapters.add(chapter)
	# Write headers
	overview_headers = ["Chapter"] + models
	for col_idx, header in enumerate(overview_headers, 1):
		cell = ws_overview.cell(row=1, column=col_idx, value=header)
		cell.fill = header_fill
		cell.font = header_font
	# Write chapter rows
	row_num = 2
	for chapter in sorted(all_chapters):
		ws_overview.cell(row=row_num, column=1, value=chapter)
		for col_idx, model in enumerate(models, 2):
			count = sum(
				1 for ch, _tp in all_results[model].values()
				if ch == chapter
			)
			ws_overview.cell(row=row_num, column=col_idx, value=count)
		row_num += 1
	# Total row
	total_cell = ws_overview.cell(row=row_num, column=1, value="TOTAL")
	total_cell.font = bold_font
	for col_idx, model in enumerate(models, 2):
		cell = ws_overview.cell(row=row_num, column=col_idx, value=len(all_results[model]))
		cell.font = bold_font
	auto_size_columns(ws_overview)

	# --- Sheet 2: Agreement Summary ---
	ws_summary = wb.create_sheet("Agreement Summary")
	total_scripts = len(get_all_scripts(all_results))
	blue_fill = openpyxl.styles.PatternFill(start_color="6699FF", fill_type="solid")
	summary_rows = [
		("Total unique scripts", total_scripts, None),
		("Full agreement (chapter + topic)", len(disagreements["topic_agree"]), green_fill),
		("Known overlaps (all valid)", len(disagreements["known_overlap"]), blue_fill),
		("Chapter disagreements", len(disagreements["chapter_disagree"]), red_fill),
		("Topic disagreements (same chapter)", len(disagreements["topic_disagree"]), yellow_fill),
		("Not in all models", len(disagreements["unique_scripts"]), cyan_fill),
	]
	# Headers
	for col_idx, header in enumerate(["Metric", "Count"], 1):
		cell = ws_summary.cell(row=1, column=col_idx, value=header)
		cell.fill = header_fill
		cell.font = header_font
	for row_idx, (label, count, fill) in enumerate(summary_rows, 2):
		ws_summary.cell(row=row_idx, column=1, value=label)
		count_cell = ws_summary.cell(row=row_idx, column=2, value=count)
		if fill:
			count_cell.fill = fill
	auto_size_columns(ws_summary)

	# --- Sheet 3: Agreements ---
	ws_agree = wb.create_sheet("Agreements")
	agree_headers = ["Folder", "Script", "Chapter/Topic"]
	for col_idx, header in enumerate(agree_headers, 1):
		cell = ws_agree.cell(row=1, column=col_idx, value=header)
		cell.fill = header_fill
		cell.font = header_font
	for row_idx, item in enumerate(disagreements["topic_agree"], 2):
		folder, basename = split_script_path(item["script"])
		ws_agree.cell(row=row_idx, column=1, value=folder)
		ws_agree.cell(row=row_idx, column=2, value=basename)
		cell_value = format_assignment(item["chapter"], item["topic"], topic_lookup)
		cell = ws_agree.cell(row=row_idx, column=3, value=cell_value)
		cell.fill = green_fill
	auto_size_columns(ws_agree)

	# --- Sheet 4: Known Overlaps ---
	ws_overlaps = wb.create_sheet("Known Overlaps")
	overlap_headers = ["Folder", "Script", "Valid Assignments"]
	for col_idx, header in enumerate(overlap_headers, 1):
		cell = ws_overlaps.cell(row=1, column=col_idx, value=header)
		cell.fill = header_fill
		cell.font = header_font
	for row_idx, item in enumerate(disagreements["known_overlap"], 2):
		folder, basename = split_script_path(item["script"])
		ws_overlaps.cell(row=row_idx, column=1, value=folder)
		ws_overlaps.cell(row=row_idx, column=2, value=basename)
		# List all valid assignments
		valid_labels = []
		for chapter, topic in sorted(item["valid_assignments"]):
			valid_labels.append(format_assignment(chapter, topic, topic_lookup))
		cell = ws_overlaps.cell(row=row_idx, column=3, value=" | ".join(valid_labels))
		cell.fill = blue_fill
	auto_size_columns(ws_overlaps)

	# --- Sheet 5: Chapter Disagreements ---
	ws_chapter = wb.create_sheet("Chapter Disagreements")
	_write_disagreement_sheet(
		ws_chapter, disagreements["chapter_disagree"], models,
		header_fill, header_font, red_fill, gray_fill, topic_lookup,
	)

	# --- Sheet 4: Topic Disagreements ---
	ws_topic = wb.create_sheet("Topic Disagreements")
	_write_disagreement_sheet(
		ws_topic, disagreements["topic_disagree"], models,
		header_fill, header_font, yellow_fill, gray_fill, topic_lookup,
	)

	# --- Sheet 5: Missing Scripts ---
	ws_missing = wb.create_sheet("Missing Scripts")
	_write_disagreement_sheet(
		ws_missing, disagreements["unique_scripts"], models,
		header_fill, header_font, cyan_fill, gray_fill, topic_lookup,
	)

	wb.save(output_path)
	console.print(f"\n[bold]Wrote XLSX to[/bold] {output_path}")

#============================================
def _find_majority_value(assignments: dict) -> str:
	"""Find the most common chapter/topic value among model assignments.

	Args:
		assignments: dict mapping model to (chapter, topic) tuple

	Returns:
		most common 'chapter/topic' string
	"""
	# Count occurrences of each chapter/topic combo
	counts = {}
	for chapter, topic in assignments.values():
		key = f"{chapter}/{topic}"
		counts[key] = counts.get(key, 0) + 1
	# Return the most common value
	majority = max(counts, key=counts.get)
	return majority

#============================================
def _write_disagreement_sheet(
	ws: openpyxl.worksheet.worksheet.Worksheet,
	items: list,
	models: list,
	header_fill: openpyxl.styles.PatternFill,
	header_font: openpyxl.styles.Font,
	disagree_fill: openpyxl.styles.PatternFill,
	missing_fill: openpyxl.styles.PatternFill,
	topic_lookup: dict,
) -> None:
	"""Write a disagreement sheet with folder/script columns and colored cells.

	Args:
		ws: openpyxl worksheet to write to
		items: list of disagreement dicts with 'script' and 'assignments' keys
		models: sorted list of model names
		header_fill: fill style for header row
		header_font: font style for header row
		disagree_fill: fill style for disagreeing cells
		missing_fill: fill style for missing cells
		topic_lookup: output of build_topic_lookup()
	"""
	# Write headers
	headers = ["Folder", "Script"] + models
	for col_idx, header in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.fill = header_fill
		cell.font = header_font

	if not items:
		ws.cell(row=2, column=1, value="(none)")
		auto_size_columns(ws)
		return

	for row_idx, item in enumerate(items, 2):
		folder, basename = split_script_path(item["script"])
		ws.cell(row=row_idx, column=1, value=folder)
		ws.cell(row=row_idx, column=2, value=basename)

		# Determine majority value for highlighting outliers
		majority = _find_majority_value(item["assignments"])

		for col_idx, model in enumerate(models, 3):
			if model in item["assignments"]:
				chapter, topic = item["assignments"][model]
				# Use chapter/topic for majority comparison, but display with name
				compare_key = f"{chapter}/{topic}"
				cell_value = format_assignment(chapter, topic, topic_lookup)
				cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
				# Highlight cells that differ from majority
				if compare_key != majority:
					cell.fill = disagree_fill
			else:
				cell = ws.cell(row=row_idx, column=col_idx, value="missing")
				cell.fill = missing_fill

	auto_size_columns(ws)

#============================================
def _to_bp_root(script: str) -> str:
	"""Normalize a script path for run_bbq_tasks.py consumption.

	Args:
		script: path from a results-*/ CSV, typically 'problems/<chapter>-problems/foo.py'

	Returns:
		path with leading 'problems/' replaced by '{bp_root}/'
	"""
	# Strip leading './' for paths like './problems/foo.py'
	clean = script.strip()
	if clean.startswith("./"):
		clean = clean[2:]
	# Already normalized - return unchanged
	if clean.startswith("{bp_root}/"):
		return clean
	# Rewrite exact 'problems/' prefix
	if clean.startswith("problems/"):
		return "{bp_root}/" + clean[len("problems/"):]
	# Anything else is unexpected - fail loudly
	raise ValueError(f"unexpected script path: {script!r}")

#============================================
def _subject_filename(subject: str) -> str:
	"""Build a deterministic per-subject CSV filename.

	Args:
		subject: subject key like 'biochemistry'

	Returns:
		filename like 'biochemistry_tasks.csv'
	"""
	key = subject.strip().lower().replace(" ", "_").replace("-", "_")
	if not key:
		raise ValueError("empty subject name")
	return f"{key}_tasks.csv"

#============================================
def _collect_agreed_rows(disagreements: dict) -> list:
	"""Build a deduplicated, sorted list of agreed task rows.

	Args:
		disagreements: output of find_disagreements()

	Returns:
		list of dict rows with keys subject, topic, script, flags, input, notes.
		Rows are unique on (subject, topic, script) and sorted by that same key.
	"""
	# Track unique (subject, topic, normalized_script) triples.
	# Note: upstream find_disagreements() still uses the legacy key name
	# 'chapter' for the biochemistry/genetics/etc. level; we translate to
	# 'subject' only in this output pathway.
	triples = set()

	# Full agreements - one row each
	for item in disagreements["topic_agree"]:
		if "chapter" not in item or "topic" not in item or "script" not in item:
			raise KeyError(f"topic_agree item missing required keys: {item!r}")
		triples.add((
			item["chapter"],
			item["topic"],
			_to_bp_root(item["script"]),
		))

	# Known overlaps - one row per valid (subject, topic) pair
	for item in disagreements["known_overlap"]:
		if "script" not in item or "valid_assignments" not in item:
			raise KeyError(f"known_overlap item missing required keys: {item!r}")
		normalized = _to_bp_root(item["script"])
		# Coerce to sorted list of tuples for deterministic ordering
		pairs = sorted(list(item["valid_assignments"]))
		for subject, topic in pairs:
			triples.add((subject, topic, normalized))

	# Convert to dict rows and sort deterministically
	rows = []
	for subject, topic, script in sorted(triples):
		rows.append({
			"subject": subject,
			"topic": topic,
			"script": script,
			"flags": "",
			"input": "",
			"notes": "",
		})
	return rows

#============================================
def write_agreed_task_csvs(disagreements: dict, output_dir: str) -> None:
	"""Write per-subject agreed_tasks CSVs for run_bbq_tasks.py.

	Args:
		disagreements: output of find_disagreements()
		output_dir: directory to write <subject>_tasks.csv files into
	"""
	os.makedirs(output_dir, exist_ok=True)
	rows = _collect_agreed_rows(disagreements)

	# Group by subject, preserving the global sort within each group
	grouped = {}
	for row in rows:
		grouped.setdefault(row["subject"], []).append(row)

	fieldnames = ["subject", "topic", "script", "flags", "input", "notes"]
	total = 0
	file_count = 0
	# Iterate subjects in sorted order for deterministic file-write order
	for subject in sorted(grouped.keys()):
		filename = _subject_filename(subject)
		out_path = os.path.join(output_dir, filename)
		subject_rows = grouped[subject]
		with open(out_path, "w", newline="") as f:
			writer = csv.DictWriter(f, fieldnames=fieldnames)
			writer.writeheader()
			for row in subject_rows:
				writer.writerow(row)
		n = len(subject_rows)
		total += n
		file_count += 1
		console.print(f"wrote {filename}: {n} rows")
	console.print(f"wrote {file_count} subject CSVs, {total} rows total -> {output_dir}")

#============================================
def main():
	"""Main entry point."""
	args = parse_args()

	# Default base directory is the script's own directory
	if args.base_dir:
		base_dir = args.base_dir
	else:
		base_dir = os.path.dirname(os.path.abspath(__file__))

	# Default output XLSX
	if args.output_xlsx:
		output_xlsx = args.output_xlsx
	else:
		output_xlsx = os.path.join(base_dir, "output", "comparison_results.xlsx")

	console.print("[bold]Loading topic indexes...[/bold]")
	topic_lookup = build_topic_lookup()

	console.print("[bold]Loading results directories...[/bold]")
	all_results = load_all_results(base_dir)

	if len(all_results) < 2:
		console.print("[red]Need at least 2 results directories to compare.[/red]")
		return

	models = sorted(all_results.keys())
	all_scripts = get_all_scripts(all_results)

	# Print overview
	print_overview_table(all_results)

	# Load known overlaps
	overlaps_path = os.path.join(base_dir, "known_overlaps.csv")
	known_overlaps = load_known_overlaps(overlaps_path)
	if known_overlaps:
		console.print(f"  Loaded [bold]{len(known_overlaps)}[/bold] known overlap scripts")

	# Find and report disagreements
	disagreements = find_disagreements(all_results, known_overlaps)
	print_agreement_summary(disagreements, len(all_scripts))
	print_chapter_disagreements(disagreements, models, topic_lookup)
	print_topic_disagreements(disagreements, models, topic_lookup)
	print_unique_scripts(disagreements, models, topic_lookup)

	# Write XLSX
	write_xlsx(all_results, disagreements, output_xlsx, topic_lookup)

	# Write per-subject agreed_tasks CSVs for run_bbq_tasks.py.
	# Default to CWD so outputs land next to wherever the user invoked the
	# tool, rather than inside the topic_classifier package directory.
	csv_dir = args.csv_output_dir or os.path.join(os.getcwd(), "agreed_tasks")
	write_agreed_task_csvs(disagreements, csv_dir)

#============================================
if __name__ == '__main__':
	main()
